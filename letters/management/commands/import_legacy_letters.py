import openpyxl
from datetime import datetime, date
from django.core.management.base import BaseCommand, CommandError
from django.utils import timezone
from letters.models import Letter, SECTOR_CHOICES, STATUS_CHOICES


class Command(BaseCommand):
    help = 'Import legacy letters from an Excel file, creating a proper audit trail entry for each row.'

    def add_arguments(self, parser):
        parser.add_argument('excel_path', type=str, help='Path to the Excel (.xlsx) file')
        parser.add_argument('--sheet', type=str, default=None, help='Sheet name (defaults to first sheet)')
        parser.add_argument('--dry-run', action='store_true', help='Preview what would be imported, without saving anything')

    def handle(self, *args, **options):
        excel_path = options['excel_path']
        dry_run = options['dry_run']

        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
        except FileNotFoundError:
            raise CommandError(f"File not found: {excel_path}")
        except Exception as e:
            raise CommandError(f"Could not open Excel file: {e}")

        sheet_name = options['sheet'] or wb.sheetnames[0]
        if sheet_name not in wb.sheetnames:
            raise CommandError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
        ws = wb[sheet_name]

        headers = [cell.value for cell in ws[1]]
        if 'serial_number' not in headers:
            raise CommandError("Excel file must have a 'serial_number' column.")

        valid_sectors = {choice[0] for choice in SECTOR_CHOICES}
        valid_statuses = {choice[0] for choice in STATUS_CHOICES}

        created_count = 0
        skipped_blank = 0
        skipped_duplicate = 0
        error_count = 0

        for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            row_data = dict(zip(headers, row))

            serial_number = row_data.get('serial_number')

            # Skip fully blank rows (common in exported Excel sheets)
            if serial_number is None:
                skipped_blank += 1
                continue

            if Letter.objects.filter(serial_number=serial_number).exists():
                self.stdout.write(self.style.WARNING(
                    f"Row {row_num}: serial_number {serial_number} already exists in DB, skipping."
                ))
                skipped_duplicate += 1
                continue

            target_sector = row_data.get('target_sector')
            if target_sector and target_sector not in valid_sectors:
                self.stdout.write(self.style.WARNING(
                    f"Row {row_num}: unrecognised sector '{target_sector}' - importing anyway, please verify."
                ))

            status = row_data.get('status') or 'PENDING'
            if status not in valid_statuses:
                self.stdout.write(self.style.WARNING(
                    f"Row {row_num}: unrecognised status '{status}', defaulting to PENDING."
                ))
                status = 'PENDING'

            date_received = self.parse_date(row_data.get('date_received'))
            replied_at = self.parse_date(row_data.get('replied_at'))
            created_at = self.parse_datetime(row_data.get('created_at'))

            if dry_run:
                self.stdout.write(
                    f"[DRY RUN] Row {row_num}: would create serial={serial_number}, "
                    f"sector={target_sector}, status={status}, created_at={created_at}"
                )
                created_count += 1
                continue

            try:
                letter = Letter(
                    serial_number=serial_number,
                    date_received=date_received,
                    sender_details=row_data.get('sender_details'),
                    letter_type=row_data.get('letter_type'),
                    accepting_officer_id=row_data.get('accepting_officer_id'),
                    target_sector=target_sector,
                    administrated_by=row_data.get('administrated_by'),
                    status=status,
                    replied_at=replied_at,
                    created_by=row_data.get('created_by') or 'LEGACY_IMPORT',
                    updated_by=row_data.get('updated_by'),
                )
                # save() triggers simple_history's post_save signal -> creates
                # a "+" (Created) HistoricalLetter row automatically.
                letter.save()

                # created_at has auto_now_add=True, so save() always overwrites it
                # with "now". Force the real legacy timestamp directly in the DB.
                if created_at:
                    Letter.objects.filter(pk=letter.pk).update(created_at=created_at)

                # Align the audit trail entry with the legacy date/reason instead
                # of showing "just now" with no context.
                history_record = letter.history.first()
                if history_record:
                    if created_at:
                        history_record.history_date = created_at
                        history_record.created_at = created_at
                    history_record.history_change_reason = "Legacy data import"
                    history_record.save()

                created_count += 1

            except Exception as e:
                self.stderr.write(self.style.ERROR(f"Row {row_num}: Error - {e}"))
                error_count += 1

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.SUCCESS(
                f"[DRY RUN] Would import: {created_count}, "
                f"Blank rows skipped: {skipped_blank}, "
                f"Already-existing skipped: {skipped_duplicate}"
            ))
        else:
            self.stdout.write(self.style.SUCCESS(
                f"Import complete. Created: {created_count}, "
                f"Blank rows skipped: {skipped_blank}, "
                f"Already-existing skipped: {skipped_duplicate}, "
                f"Errors: {error_count}"
            ))

    def parse_date(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, str):
            value = value.strip()
            for fmt in ('%Y.%m.%d', '%Y-%m-%d', '%d/%m/%Y', '%d.%m.%Y'):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    def parse_datetime(self, value):
        if not value:
            return None
        if isinstance(value, datetime):
            if timezone.is_naive(value):
                return timezone.make_aware(value)
            return value
        if isinstance(value, str):
            value = value.strip()
            for fmt in ('%Y-%m-%d %H:%M:%S.%f%z', '%Y-%m-%d %H:%M:%S%z', '%Y-%m-%d %H:%M:%S'):
                try:
                    dt = datetime.strptime(value, fmt)
                    if timezone.is_naive(dt):
                        dt = timezone.make_aware(dt)
                    return dt
                except ValueError:
                    continue
        return None