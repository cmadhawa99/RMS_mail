from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from .models import Letter, SectorProfile
from django.views.decorators.cache import never_cache  # <--- CRITICAL SECURITY TOOL
from django.core.paginator import Paginator
from .forms import UserForm, LetterForm
from django.contrib.auth.models import User
from django.contrib import messages
from django.contrib.auth import logout  # <--- Needed for logout

import openpyxl
from openpyxl import Workbook
from django.http import HttpResponse
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- PUBLIC PORTAL ---

@never_cache  # Security: Prevents "Back" button from showing this after logout
@login_required
def sector_dashboard(request):
    # 1) Redirect Superusers to Admin Panel
    if request.user.is_superuser:
        return redirect('custom_admin_dashboard')

    try:
        user_profile = request.user.sectorprofile
        user_sector = user_profile.sector
    except SectorProfile.DoesNotExist:
        if request.user.is_superuser:
            user_sector = "ADMIN"
        else:
            user_sector = "NONE"


    letters = Letter.objects.all().order_by('-date_received')

    selected_sector = request.GET.get('sector', 'ALL')

    if selected_sector != "ALL":
        letters = letters.filter(target_sector=selected_sector)

    search_query = request.GET.get('q', '')
    search_type = request.GET.get('search_type', 'all')

    if search_query:
        if search_type == 'serial':
            letters = letters.filter(serial_number__iexact=search_query)
        elif search_type == 'date':
            letters = letters.filter(date_received__icontains=search_query)
        else:
            letters = letters.filter(
            Q(serial_number__icontains=search_query) |
            Q(sender_name__icontains=search_query) |
            Q(letter_type__icontains=search_query)

        )

    total_count = letters.count()
    resolved_count = letters.filter(is_replied=True).count()
    pending_count = total_count - resolved_count

    paginator = Paginator(letters, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    if request.method == 'POST':
        letter_id = request.POST.get('letter_id')
        reply_date = request.POST.get('reply_date')

        letter = get_object_or_404(Letter, pk=letter_id, target_sector=user_sector) #check for problems this line

        #Updated part 2026-01-28 3.14PM
        #Security: Ensure user owns this sector
        if letter.target_sector != user_sector:
            messages.error(request, "Access Denied: You cannot update letters from other sectors.")
            return redirect(f"{request.path}?sector={selected_sector}&page={page_obj.number}")

        # Update Text Fields
        if reply_date:
            letter.replied_at = reply_date
            letter.is_replied = True

        # Handle File Uploads (Slots 1-6)
        # Iterate through attachment_1 to attachment_6
        files_uploaded = False
        for i in range(1, 7):
            field_name = f'attachment_{i}'
            #Check if a file was sent for this specific slot
            if request.FILES.get(field_name):
                setattr(letter, field_name, request.FILES.get(field_name))
                files_uploaded = True
        letter.save()

        #Success Message
        msg = f"letter #{letter.serial_number} updated successfully"
        if files_uploaded:
            msg += " (Documents uploaded)"

        messages.success(request, msg)

        return redirect(f"{request.path}?sector={selected_sector}&page={page_obj.number}")

    context = {
        'user_sector': user_sector,
        'selected_sector': selected_sector,
        'search_query': search_query,
        'search_type': search_type,
        'letters': page_obj,
        'total': total_count,
        'pending': pending_count,
        'resolved': resolved_count,
    }

    return render(request, 'letters/user/dashboard.html', context)


@never_cache
@login_required
def letter_detail(request, pk):
    letter = get_object_or_404(Letter, pk=pk)

    try:
        user_profile = request.user.sectorprofile
        user_sector = user_profile.sector
    except SectorProfile.DoesNotExist:
        if request.user.is_superuser:
            user_sector = "ADMIN"
        else:
            user_sector = "NONE"

    if user_sector != "ADMIN" and letter.target_sector != user_sector:
        return render(request, 'letters/common/access_denied.html')

    return render(request, 'letters/user/letter_detail.html', {
        'letter': letter,
        'user_sector': user_sector
    })

@login_required
def view_letter_images(request, pk):
    letter = get_object_or_404(Letter, pk=pk)

    try:
        user_profile =  request.user.sectorprofile
        user_sector = user_profile.sector
    except SectorProfile.DoesNotExist:
        user_sector = "NONE"

        if not request.user.is_superuser:
            if user_sector != "ADMIN" and letter.target_sector != user_sector:
                return render(request, 'letters/common/access_denied.html')

    attachments = []
    if letter.attachment_1: attachments.append(letter.attachment_1)
    if letter.attachment_2: attachments.append(letter.attachment_2)
    if letter.attachment_3: attachments.append(letter.attachment_3)
    if letter.attachment_4: attachments.append(letter.attachment_4)
    if letter.attachment_5: attachments.append(letter.attachment_5)
    if letter.attachment_6: attachments.append(letter.attachment_6)

    return render(request, 'letters/common/letter_images.html', {
        'letter': letter,
        'attachments': attachments,
    })




# ------- CUSTOM ADMIN PANEL VIEWS -------

@never_cache
@login_required
def custom_admin_dashboard(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    return render(request, 'letters/admin/pages/admin_dashboard.html')


@never_cache
@login_required
def custom_admin_users(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    users = User.objects.filter(is_superuser=False).select_related('sectorprofile')

    search_query = request.GET.get('q', '')
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )

    return render(request, 'letters/admin/pages/admin_users.html', {'users': users, 'search_query': search_query})


@never_cache
@login_required
def custom_admin_letters(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    letters_list = Letter.objects.all().order_by('-date_received')

    search_query = request.GET.get('q', '')
    search_type = request.GET.get('search_type', 'all')
    if search_query:
        if search_type == 'serial':
            letters_list = letters_list.filter(serial_number__iexact=search_query)
        elif search_type == 'date':
            letters_list = letters_list.filter(date_received__icontains=search_query)
        else:
            letters_list = letters_list.filter(
            Q(serial_number__icontains=search_query) |
            Q(sender_name__icontains=search_query) |
            Q(letter_type__icontains=search_query) |
            Q(target_sector__icontains=search_query)
        )

    paginator = Paginator(letters_list, 20)
    page_number = request.GET.get('page')
    letters = paginator.get_page(page_number)
    return render(request, 'letters/admin/pages/admin_letters.html', {'letters': letters, 'search_query': search_query})


# --- ADMIN DETAILS (With Security) ---

@never_cache
@login_required
def admin_user_detail(request, user_id):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    user_obj = get_object_or_404(User, pk=user_id)
    return render(request, 'letters/admin/pages/admin_user_detail.html', {'user_obj': user_obj})


@never_cache
@login_required
def admin_letter_detail(request, pk):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    letter = get_object_or_404(Letter, pk=pk)
    return render(request, 'letters/admin/pages/admin_letter_detail.html', {'letter': letter})


# --- ACTION VIEWS ---

@never_cache
@login_required
def create_user(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "New officer account created successfully!")
            return redirect('custom_admin_users')
    else:
        form = UserForm()

    return render(request, 'letters/admin/pages/user_form.html', {'form': form, 'title': 'Create New Officer'})


@never_cache
@login_required
def edit_user(request, user_id):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    user_obj = get_object_or_404(User, pk=user_id)

    if request.method == 'POST':
        form = UserForm(request.POST, instance=user_obj)
        if form.is_valid():
            form.save()
            messages.success(request, f"Details for {user_obj.username} updated.")
            return redirect('admin_user_detail', user_id=user_obj.id)
    else:
        form = UserForm(instance=user_obj)

    return render(request, 'letters/admin/pages/user_form.html', {'form': form, 'title': 'Edit Officer Details'})


@login_required
def delete_user(request, user_id):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    if request.method == 'POST':
        user_to_delete = get_object_or_404(User, pk=user_id)
        if not user_to_delete.is_superuser:
            user_to_delete.delete()
            messages.success(request, "User account deleted.")

    return redirect('custom_admin_users')


@never_cache
@login_required
def add_letter(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    if request.method == 'POST':
        form = LetterForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()

            messages.success(request, "New letter and pages added successfully.")
            return redirect('custom_admin_letters')
    else:
        form = LetterForm()

    return render(request, 'letters/admin/pages/letter_form.html', {'form': form, 'title': 'Add New Letter'})


@never_cache
@login_required
def edit_letter(request, pk):
    if not request.user.is_superuser: return redirect('sector_dashboard')
    letter = get_object_or_404(Letter, pk=pk)

    if request.method == 'POST':
        form = LetterForm(request.POST, request.FILES, instance=letter)
        if form.is_valid():
            form.save()
            messages.success(request, "Letter updated successfully.")
            return redirect('admin_letter_detail', pk=letter.pk)
    else:
        form = LetterForm(instance=letter)

    return render(request, 'letters/admin/pages/letter_form.html', {'form': form, 'title': 'Edit Letter'})


@login_required
def delete_letter(request, pk):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    letter = get_object_or_404(Letter, pk=pk)
    if request.method == 'POST':
        letter.delete()
        messages.success(request, "Letter permanently deleted.")

    return redirect('custom_admin_letters')


# --- LOGOUT VIEW (NEW) ---

def logout_view(request):
    logout(request)
    messages.info(request, "You have been logged out.")
    return redirect('login')


@never_cache
@login_required
def export_letters_excel(request):
    if not request.user.is_superuser: return redirect('sector_dashboard')

    #1) Get the data (Apply search filter if it exists)
    letters = Letter.objects.all().order_by('-date_received')
    search_query = request.GET.get('q', '')

    if search_query:
        letters = letters.filter(
            Q(serial_number__icontains=search_query) |
            Q(sender_name__icontains=search_query) |
            Q(letter_type__icontains=search_query) |
            Q(target_sector__icontains=search_query)
        )

        clean_query = "".join([c for c in search_query if c.isalnum() or c in (' ', '-', '_')]).strip()
        filename = f"Search_Results_{clean_query}.xlsx"

    else:
        year = datetime.now().year
        filename = f"Weligepola_Pradeshiya_sabha_letters_database_{year}.xlsx"


    # 2) Create the Excel Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Letters Data"

    # --- Styles added 2026-01-29 ---

    sector_colors = {
        'GOVERNING': 'd99694',
        'HEALTH': 'b3a2c7',
        'DEVELOPMENT': 'e46c0a',
        'INCOME': 'ffff00',
        'ACCOUNTS': '95b3d7'
    }

    sector_map = {
        'GOVERNING': 'පාලන අංශය',
        'HEALTH': 'සෞඛ්‍ය අංශය',
        'DEVELOPMENT': 'සංවර්ධන අංශය',
        'INCOME': 'ආදායම් අංශය',
        'ACCOUNTS': 'ගිණුම් අංශය',
    }

    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin') )

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color='2b2b2b', end_color='2b2b2b', fill_type='solid')

    title_font = Font(bold=True, color="1f4e78", size=16)

    # 3) Add the Title
    ws.merge_cells('A1:I1')
    title_cell = ws['A1']
    title_cell.value = "ලිපි ලේඛන කළමනාකරණ පද්ධතිය - වැලිගෙපොල ප්‍රාදේශීය සභාව"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    # 4) Add Color Legend
    ws['A3'] = "වර්ණ දර්ශකය (Color legend)"
    ws['A3'].font = Font(bold=True, underline="single")

    current_row = 4
    for sector_code, color_hex in sector_colors.items():
        label_cell = ws.cell(row=current_row, column=1)
        label_cell.value = sector_map.get(sector_code, sector_code)
        label_cell.border = thin_border

        color_cell = ws.cell(row=current_row, column=2)
        color_cell.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')
        color_cell.border = thin_border

        current_row += 1
    start_row = current_row + 2


    # 5) Add Headers (In Sinhala)
    headers = [
        'ලිපි අංකය (Serial)',
        'ලැබුණු දිනය (Date)',
        'එවූ අයගේ නම (Sender)',
        'ලිපියේ වර්ගය (Type)',
        'අංශය (Sector)',
        'පරිපාලනය කළේ (Admin By)',
        'භාරගත් නිලධාරී (Accepting Officer)',
        'තත්වය (Status)',
        'පිළිතුරු දුන් දිනය (Replied Date)'

    ]

    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # 6) Add Rows
    for letter in letters:
        start_row += 1

        #Translate status
        status = "පිළිතුරු යොමු කර ඇත" if letter.is_replied else "විමර්ශනය වෙමින් පවතී "

        #Translate Sector
        sector_sinhala = sector_map.get(letter.target_sector, letter.target_sector)

        #Handle Admin By
        admin_by = letter.get_administrated_by_display()

        #Handle date formatting
        replied_at = ""
        if letter.replied_at:
            replied_at = letter.replied_at.strftime("%Y-%m-%d")

        row_data = [
            letter.serial_number,
            letter.date_received,
            letter.sender_name,
            letter.letter_type,
            sector_sinhala,
            admin_by,
            letter.accepting_officer_id,
            status,
            replied_at
        ]

        row_color = sector_colors.get(letter.target_sector, "FFFFFF")
        row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type='solid')

        for col_num, cell_value in enumerate(row_data, 1):
            cell = ws.cell(row=start_row, column=col_num)
            cell.value = cell_value
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

            if col_num in  [1, 2, 5, 8, 9]:
                cell.alignment = Alignment(horizontal='center', vertical='center')


    # 7) Auto-adjust column width
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:

            if cell.row < 9:
                continue

            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = min(max_length + 3, 50)
        ws.column_dimensions[column].width = adjusted_width


    # 6) Prepare the response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    wb.save(response)
    return response
#test